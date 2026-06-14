import json
import os
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from src.patcher.core.data_manager import DataManager
from src.patcher.core.exceptions import PatcherError
from src.patcher.core.exporter import Exporter
from src.patcher.core.models.patch import PatchDevice, PatchTitle


def _frame(titles):
    """The canonical DataFrame the way DataManager hands it to the exporter."""
    return DataManager(disable_cache=True)._create_dataframe(titles)


class TestSerialize:
    def test_serialize_titles_to_dict_shape(self, sample_patch_reports):
        """The serializer wraps titles in metadata and preserves model fields."""
        payload = Exporter(sample_patch_reports).serialize_titles_to_dict(
            report_title="Test Report"
        )

        assert payload["report_title"] == "Test Report"
        assert payload["title_count"] == 1
        assert "generated_at" in payload  # ISO timestamp
        assert isinstance(payload["titles"], list)

        title = payload["titles"][0]
        assert title["title"] == "Example Software"
        assert title["title_id"] == "0"
        assert title["completion_percent"] == 83.33
        assert title["total_hosts"] == 12

    def test_serialize_titles_to_dict_empty(self):
        """An empty title list still produces a valid envelope."""
        payload = Exporter([]).serialize_titles_to_dict(report_title=None)
        assert payload["title_count"] == 0
        assert payload["titles"] == []
        assert payload["report_title"] is None

    def test_serialize_titles_to_dict_is_json_serializable(self, sample_patch_reports):
        """The dict must round-trip through json.dumps without errors."""
        payload = Exporter(sample_patch_reports).serialize_titles_to_dict(report_title="Test")
        parsed = json.loads(json.dumps(payload))
        assert parsed["title_count"] == 1


class TestExportExcel:
    def test_write_multisheet_workbook_creates_per_title_sheets(self, tmp_path):
        """The --device-details Excel writer adds a per-title sheet and skips empty ones."""
        titles = [
            PatchTitle(
                title="Firefox",
                title_id="1",
                released="2026-01-01",
                hosts_patched=10,
                missing_patch=2,
                latest_version="120.0",
            )
        ]
        device = PatchDevice(
            computer_name="Mac-01",
            device_id="1",
            username="jappleseed",
            operating_system_version="14.5",
            last_contact_time=datetime(2026, 1, 1, 12, 0, 0),
            version="119.0",
        )
        device_reports = {"1": [device], "2": []}  # title "2" has no devices → no sheet
        df = pd.DataFrame([{"Title": "Firefox", "Hosts": 10}])

        out = tmp_path / "report.xlsx"
        Exporter(titles)._write_multisheet_workbook(out, df, device_reports)

        wb = load_workbook(out)
        assert wb.sheetnames == ["Patch Report", "Firefox"]  # empty title skipped
        ws = wb["Firefox"]
        assert ws["A1"].value == "Computer Name"
        assert ws["A2"].value == "Mac-01"
        assert ws["B2"].value == "1"  # device_id rendered as string

    @pytest.mark.asyncio
    async def test_export_to_excel_success(self, sample_patch_reports, temp_output_dir):
        df = _frame(sample_patch_reports)
        exported = await Exporter(sample_patch_reports).export(
            df, temp_output_dir, "Test Report", formats={"excel"}
        )

        excel_path = exported.get("excel")
        assert excel_path is not None
        assert os.path.exists(excel_path)

        out = pd.read_excel(excel_path)
        assert not out.empty
        assert list(out.columns) == [
            "Title",
            "Released",
            "Hosts Patched",
            "Missing Patch",
            "Latest Version",
            "Completion Percent",
            "Total Hosts",
        ]

    @pytest.mark.asyncio
    async def test_export_never_adds_integration_columns_even_when_matched(self, temp_output_dir):
        """Rendered reports carry no integration coverage column, even when a cask matched."""
        from src.patcher.core.models.cask import CaskMatch

        matched = PatchTitle(
            title="Firefox",
            title_id="1",
            released="2026-01-01",
            hosts_patched=5,
            missing_patch=5,
            latest_version="1.0",
            homebrew_cask=[CaskMatch(name="Firefox", token="firefox", version="1.0")],
        )
        unmatched = PatchTitle(
            title="Acme",
            title_id="2",
            released="2026-01-01",
            hosts_patched=1,
            missing_patch=1,
            latest_version="2.0",
        )
        titles = [matched, unmatched]
        df = _frame(titles)

        exported = await Exporter(titles).export(
            df, temp_output_dir, "Test Report", formats={"excel"}
        )
        out = pd.read_excel(exported["excel"])

        # The cask match still rides on the title (it reaches analyze/JSON), but a
        # rendered report surfaces no integration column, derived or raw.
        assert "Homebrew" not in out.columns
        assert "Homebrew Cask" not in out.columns

    @pytest.mark.asyncio
    async def test_export_omits_homebrew_column_without_matches(
        self, sample_patch_reports, temp_output_dir
    ):
        """Installomator-only exports gain no Homebrew column either."""
        df = _frame(sample_patch_reports)
        exported = await Exporter(sample_patch_reports).export(
            df, temp_output_dir, "Test Report", formats={"excel"}
        )
        out = pd.read_excel(exported["excel"])

        assert "Homebrew" not in out.columns

    @pytest.mark.asyncio
    async def test_export_to_excel_invalid_directory(self, sample_patch_reports):
        """An unwritable output directory surfaces a PatcherError."""
        df = _frame(sample_patch_reports)
        with patch.object(Path, "mkdir", side_effect=OSError("Test Invalid Directory Error")):
            with pytest.raises(PatcherError, match="Encountered error saving DataFrame"):
                await Exporter(sample_patch_reports).export(
                    df, "/invalid/path/to/output", "Test Report", formats={"excel"}
                )

    @pytest.mark.asyncio
    async def test_export_to_excel_permission_error(self, sample_patch_reports, temp_output_path):
        """A permission error while preparing the output directory surfaces a PatcherError."""
        df = _frame(sample_patch_reports)
        temp_file = temp_output_path / "patch-report.xlsx"
        with patch.object(Path, "mkdir", side_effect=PermissionError("Test Permission Error")):
            with pytest.raises(PatcherError, match="Encountered error saving DataFrame"):
                await Exporter(sample_patch_reports).export(
                    df, temp_file, "Test Report", formats={"excel"}
                )


class TestExportJson:
    @pytest.mark.asyncio
    async def test_export_to_json_success(self, sample_patch_reports, temp_output_dir):
        """The export method writes a valid JSON file when 'json' is in formats."""
        df = _frame(sample_patch_reports)
        exported = await Exporter(sample_patch_reports).export(
            df, temp_output_dir, "Test Report", formats={"json"}
        )

        json_path = exported.get("json")
        assert json_path is not None
        assert os.path.exists(json_path)
        assert json_path.endswith(".json")

        with open(json_path) as f:
            payload = json.load(f)

        assert payload["title_count"] == 1
        assert payload["report_title"] == "Test Report"
        assert payload["titles"][0]["title"] == "Example Software"
        assert "generated_at" in payload

    @pytest.mark.asyncio
    async def test_export_default_includes_all_formats(self, sample_patch_reports, temp_output_dir):
        """When no formats are specified, all four are emitted."""
        df = _frame(sample_patch_reports)
        exported = await Exporter(sample_patch_reports).export(df, temp_output_dir, "Test Report")

        assert "json" in exported
        assert "excel" in exported
        assert "html" in exported
        assert "pdf" in exported


class TestExportPdf:
    @pytest.mark.asyncio
    async def test_export_pdf_threads_ui_config_to_pdf_report(self, monkeypatch, tmp_path):
        """
        Regression for #69: the exporter must pass its ``ui_config`` into the
        ``PDFReport`` it constructs. Before the fix, ``_export_pdf`` did
        ``PDFReport(date_format=date_format)`` with no ``ui_config``, so the PDF
        fell through to :class:`UIDefaults` placeholders ("Default header text"
        / "Default footer text") regardless of what the user had configured.
        """
        captured: dict = {}

        def fake_pdf_report(date_format, ui_config=None):
            captured["date_format"] = date_format
            captured["ui_config"] = ui_config
            # Minimal stand-in for the PDFReport surface ``_export_pdf`` exercises.
            pdf = MagicMock()
            pdf.ui_config = ui_config or {}
            pdf.calculate_column_widths = lambda df: [10] * len(df.columns)
            pdf.h = 200
            pdf.get_y = lambda: 0
            return pdf

        monkeypatch.setattr("src.patcher.core.exporter.PDFReport", fake_pdf_report)

        custom_ui = {
            "header_text": "May-Patch-Report-2026",
            "footer_text": "Confidential",
            "font_name": "Helvetica",
            "reg_font_path": "",
            "bold_font_path": "",
            "logo_path": "",
            "header_color": "#abcdef",
        }
        exporter = Exporter([], ui_config=custom_ui)
        df = pd.DataFrame({"Title": ["Firefox"], "Latest Version": ["121.0"]})
        pdf_path = tmp_path / "test.pdf"

        await exporter._export_pdf(df, pdf_path, "%Y-%m-%d")

        # The fix: ``ui_config`` reaches ``PDFReport`` so the configured header
        # text actually renders, instead of the ``UIDefaults`` placeholder.
        assert captured["ui_config"] is custom_ui
        assert captured["ui_config"]["header_text"] == "May-Patch-Report-2026"
        assert captured["ui_config"]["footer_text"] == "Confidential"
