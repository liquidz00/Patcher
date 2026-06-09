import json
import os
import pickle
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal
from src.patcher.core.data_manager import DataManager, serialize_titles_to_dict
from src.patcher.core.exceptions import PatcherError
from src.patcher.core.models.patch import PatchDevice, PatchTitle


@pytest.fixture
def mock_formats():
    return {"excel"}


class TestExportExcel:
    def test_write_multisheet_workbook_creates_per_title_sheets(self, tmp_path):
        """The --device-details Excel writer adds a per-title sheet and skips empty ones."""
        dm = DataManager()
        df = pd.DataFrame([{"Title": "Firefox", "Hosts": 10}])
        titles = [
            PatchTitle(
                title="Firefox",
                title_id="1",
                released="2026-01-01",
                hosts_patched=10,
                missing_patch=2,
                latest_version="120.0",
            )
        ]
        device = PatchDevice(
            computer_name="Mac-01",
            device_id="1",
            username="jappleseed",
            operating_system_version="14.5",
            last_contact_time=datetime(2026, 1, 1, 12, 0, 0),
            version="119.0",
        )
        device_reports = {"1": [device], "2": []}  # title "2" has no devices → no sheet

        out = tmp_path / "report.xlsx"
        dm._write_multisheet_workbook(out, df, titles, device_reports)

        wb = load_workbook(out)
        assert wb.sheetnames == ["Patch Report", "Firefox"]  # empty title skipped
        ws = wb["Firefox"]
        assert ws["A1"].value == "Computer Name"
        assert ws["A2"].value == "Mac-01"
        assert ws["B2"].value == "1"  # device_id rendered as string

    @pytest.mark.asyncio
    async def test_export_to_excel_success(
        self, sample_patch_reports, temp_output_dir, mock_formats
    ):
        data_manager = DataManager()

        with patch.object(data_manager, "_cache_data", return_value=None) as mock_cache_data:
            exported_files = await data_manager.export(
                sample_patch_reports, temp_output_dir, "Test Report", formats=mock_formats
            )

            excel_path = exported_files.get("excel")  # type: ignore
            assert excel_path is not None
            assert os.path.exists(excel_path)

            df = pd.read_excel(excel_path)
            assert not df.empty
            assert list(df.columns) == [
                "Title",
                # "Title Id",
                "Released",
                "Hosts Patched",
                "Missing Patch",
                "Latest Version",
                "Completion Percent",
                "Total Hosts",
                # "Install Label",
            ]

            args, _ = mock_cache_data.call_args
            cached_df = args[0]

            # Account for cached dataframe having all columns
            common_columns = df.columns.intersection(cached_df.columns)
            assert_frame_equal(df[common_columns], cached_df[common_columns], check_like=True)

    @pytest.mark.asyncio
    async def test_export_adds_homebrew_column_when_matched(self, temp_output_dir, mock_formats):
        """A Homebrew coverage column appears (showing cask tokens) only when titles matched a cask."""
        from src.patcher.core.models.cask import CaskMatch

        matched = PatchTitle(
            title="Firefox",
            title_id="1",
            released="2026-01-01",
            hosts_patched=5,
            missing_patch=5,
            latest_version="1.0",
            homebrew_cask=[CaskMatch(name="Firefox", token="firefox", version="1.0")],
        )
        unmatched = PatchTitle(
            title="Acme",
            title_id="2",
            released="2026-01-01",
            hosts_patched=1,
            missing_patch=1,
            latest_version="2.0",
        )
        data_manager = DataManager(disable_cache=True)

        exported = await data_manager.export(
            [matched, unmatched], temp_output_dir, "Test Report", formats=mock_formats
        )
        df = pd.read_excel(exported["excel"])

        assert "Homebrew" in df.columns
        assert df.loc[df["Title"] == "Firefox", "Homebrew"].iloc[0] == "firefox"
        # Raw list-of-dicts field is never surfaced as a column.
        assert "Homebrew Cask" not in df.columns

    @pytest.mark.asyncio
    async def test_export_omits_homebrew_column_without_matches(
        self, sample_patch_reports, temp_output_dir, mock_formats
    ):
        """Default (Installomator-only) exports gain no Homebrew column."""
        data_manager = DataManager(disable_cache=True)

        exported = await data_manager.export(
            sample_patch_reports, temp_output_dir, "Test Report", formats=mock_formats
        )
        df = pd.read_excel(exported["excel"])

        assert "Homebrew" not in df.columns

    @pytest.mark.asyncio
    async def test_export_to_excel_dataframe_creation_error(
        self, mock_data_manager, temp_output_dir, mock_formats
    ):
        with patch.object(pd, "DataFrame", side_effect=ValueError("Test Error")):
            with pytest.raises(PatcherError, match="Encountered error creating DataFrame."):
                await mock_data_manager.export(
                    [], temp_output_dir, "Test Report", formats=mock_formats
                )

    # Edge case tests
    @pytest.mark.asyncio
    async def test_export_to_excel_empty_patch_reports(
        self, mock_data_manager, temp_output_dir, mock_formats
    ):
        """Ensure export_excel handles empty patch_reports gracefully."""
        with patch.object(pd, "DataFrame", side_effect=pd.errors.EmptyDataError):
            with pytest.raises(PatcherError, match="Encountered error creating DataFrame"):
                await mock_data_manager.export(
                    [], temp_output_dir, "Test Report", formats=mock_formats
                )

    @pytest.mark.asyncio
    async def test_export_to_excel_invalid_directory(self, mock_data_manager, mock_formats):
        """Ensure export_excel raises an error for invalid output directory."""
        invalid_dir = "/invalid/path/to/output"

        with patch.object(mock_data_manager, "_cache_data", return_value=None):
            with patch.object(Path, "mkdir", side_effect=OSError("Test Invalid Directory Error")):
                with pytest.raises(PatcherError, match="Encountered error saving DataFrame"):
                    await mock_data_manager.export(
                        [], invalid_dir, "Test Report", formats=mock_formats
                    )

    @pytest.mark.asyncio
    async def test_export_to_excel_permission_error(
        self, mock_data_manager, temp_output_path, mock_formats
    ):
        """Simulate a permission error when writing to an output directory."""
        temp_file = temp_output_path / "patch-report.xlsx"

        with patch.object(mock_data_manager, "_cache_data", return_value=None):
            with patch.object(Path, "mkdir", side_effect=PermissionError("Test Permission Error")):
                with pytest.raises(PatcherError, match="Encountered error saving DataFrame"):
                    await mock_data_manager.export(
                        [], temp_file, "Test Report", formats=mock_formats
                    )


class TestCache:
    def test_cache_property(self):
        data_manager = DataManager()
        assert data_manager.cache_off is False

    def test_cache_property_false(self):
        data_manager = DataManager(disable_cache=True)
        assert data_manager.cache_off is True

    def test_clean_cache_removes_expired_files(self, temp_output_path):
        """Ensure clean_cache removes only expired files."""
        data_manager = DataManager()
        data_manager.cache_dir = temp_output_path

        expired_file = temp_output_path / "expired.pkl"
        valid_file = temp_output_path / "valid.pkl"

        expired_file.touch()
        valid_file.touch()

        expired_time = (datetime.now() - timedelta(days=91)).timestamp()
        valid_time = (datetime.now() - timedelta(days=15)).timestamp()
        os.utime(expired_file, (expired_time, expired_time))
        os.utime(valid_file, (valid_time, valid_time))

        mock_iterdir = MagicMock(return_value=[expired_file, valid_file])

        with patch.object(Path, "iterdir", mock_iterdir):
            data_manager._clean_cache()

        # Ensure only the expired file is removed
        assert not expired_file.exists()
        assert valid_file.exists()

    def test_clean_cache_no_permissions(self):
        """Ensure clean_cache handles missing permissions gracefully."""
        # Mock the cache directory and files
        with (
            patch("pathlib.Path.iterdir") as mock_iterdir,
            patch("pathlib.Path.unlink", side_effect=PermissionError("Test Permission Error")),
        ):
            # Simulate a file in the cache directory
            cache_file = MagicMock()
            cache_file.suffix = ".pkl"
            cache_file.stat.return_value.st_mtime = (
                datetime.now() - timedelta(days=91)
            ).timestamp()
            mock_iterdir.return_value = [cache_file]

            # Initialize DataManager and run the method
            data_manager = DataManager(disable_cache=True)
            data_manager._clean_cache()

            # Ensure the file was attempted to be deleted, but an error was logged
            cache_file.unlink.assert_called_once()

    def test_get_latest_dataset_no_files(self):
        """Ensure get_latest_dataset returns None when no datasets are available."""
        data_manager = DataManager(disable_cache=True)
        with patch.object(data_manager, "get_cached_files", return_value=[]):
            assert data_manager.get_latest_dataset() is None

    @pytest.mark.asyncio
    async def test_load_cached_data_with_corrupted_files(self, mock_data_manager):
        """Ensure load_cached_data skips corrupted files."""
        corrupted_file = MagicMock(spec=Path)
        corrupted_file.name = "corrupted_cache.pkl"
        corrupted_file.__str__.return_value = "/mocked/path/corrupted_cache.pkl"

        with patch.object(mock_data_manager, "get_cached_files", return_value=[corrupted_file]):
            with patch("pickle.load", side_effect=pickle.UnpicklingError("Test Unpickling error")):
                loaded_data = mock_data_manager.load_cached_data()

        # Assert that no valid data was loaded
        assert len(loaded_data) == 0

    def test_cache_data_writes_parquet_and_load_round_trips(self, tmp_path):
        """New caches are written as Parquet (version-stable) and load() reads them back."""
        dm = DataManager()
        dm.cache_dir = tmp_path
        df = pd.DataFrame([{"title": "Firefox", "completion_percent": 80.0, "total_hosts": 10}])

        dm._cache_data(df)

        parquet_files = list(tmp_path.glob("*.parquet"))
        assert len(parquet_files) == 1  # Parquet, not pickle
        assert_frame_equal(DataManager.load(parquet_files[0]), df)

    def test_load_wraps_unreadable_pickle_with_recovery_hint(self, tmp_path):
        """A legacy .pkl that can't be unpickled surfaces an actionable PatcherError."""
        bad = tmp_path / "patch_data_bad.pkl"
        bad.write_bytes(b"not a valid pickle stream")
        with pytest.raises(PatcherError, match="different version of pandas") as excinfo:
            DataManager.load(bad)
        assert "reset cache" in excinfo.value.recovery  # presentation-only attr, not in str()


class TestTitlesProperty:
    def test_titles_property_uninitialized(self):
        """Test titles property access when uninitialized."""
        data_manager = DataManager()
        with patch.object(data_manager, "get_latest_dataset", return_value=None):
            with pytest.raises(
                PatcherError, match="No dataset available, unable to proceed with validation."
            ):
                _ = data_manager.titles

    def test_titles_property_setter_invalid_type(self):
        """Test titles setter with invalid type."""
        data_manager = DataManager()
        with pytest.raises(
            PatcherError, match="Value Invalid Type must be an list of PatchTitle objects."
        ):
            data_manager.titles = "Invalid Type"

    def test_titles_property_setter_empty_list(self):
        """Test titles setter with an empty list."""
        data_manager = DataManager()
        with pytest.raises(PatcherError, match="PatchTitles cannot be set to an empty list"):
            data_manager.titles = []

    def test_titles_property_setter_valid(self):
        """Test titles setter with valid PatchTitle objects."""
        data_manager = DataManager()
        patch_titles = [
            PatchTitle(
                title="Patch A",
                title_id="0",
                released="2022-01-01",
                hosts_patched=50,
                missing_patch=10,
                latest_version="1.0.0",
            ),
            PatchTitle(
                title="Patch B",
                title_id="1",
                released="2023-01-01",
                hosts_patched=30,
                missing_patch=20,
                latest_version="2.0.0",
            ),
        ]
        data_manager.titles = patch_titles
        assert data_manager.titles == patch_titles


class TestSerialize:
    def test_serialize_titles_to_dict_shape(self, sample_patch_reports):
        """The serializer wraps titles in metadata and preserves model fields."""
        payload = serialize_titles_to_dict(sample_patch_reports, report_title="Test Report")

        assert payload["report_title"] == "Test Report"
        assert payload["title_count"] == 1
        assert "generated_at" in payload  # ISO timestamp
        assert isinstance(payload["titles"], list)

        title = payload["titles"][0]
        assert title["title"] == "Example Software"
        assert title["title_id"] == "0"
        assert title["completion_percent"] == 83.33
        assert title["total_hosts"] == 12

    def test_serialize_titles_to_dict_empty(self):
        """An empty title list still produces a valid envelope."""
        payload = serialize_titles_to_dict([], report_title=None)
        assert payload["title_count"] == 0
        assert payload["titles"] == []
        assert payload["report_title"] is None

    def test_serialize_titles_to_dict_is_json_serializable(self, sample_patch_reports):
        """The dict must round-trip through json.dumps without errors."""
        payload = serialize_titles_to_dict(sample_patch_reports, report_title="Test")
        json_str = json.dumps(payload)
        parsed = json.loads(json_str)
        assert parsed["title_count"] == 1


class TestExportJson:
    @pytest.mark.asyncio
    async def test_export_to_json_success(self, sample_patch_reports, temp_output_dir):
        """The export method writes a valid JSON file when 'json' is in formats."""
        data_manager = DataManager()

        with patch.object(data_manager, "_cache_data", return_value=None):
            exported_files = await data_manager.export(
                sample_patch_reports, temp_output_dir, "Test Report", formats={"json"}
            )

        json_path = exported_files.get("json")
        assert json_path is not None
        assert os.path.exists(json_path)
        assert json_path.endswith(".json")

        with open(json_path) as f:
            payload = json.load(f)

        assert payload["title_count"] == 1
        assert payload["report_title"] == "Test Report"
        assert payload["titles"][0]["title"] == "Example Software"
        assert "generated_at" in payload

    @pytest.mark.asyncio
    async def test_export_default_includes_json(self, sample_patch_reports, temp_output_dir):
        """When no formats are specified, JSON is included in the default set."""
        data_manager = DataManager()

        with patch.object(data_manager, "_cache_data", return_value=None):
            exported_files = await data_manager.export(
                sample_patch_reports, temp_output_dir, "Test Report"
            )

        assert "json" in exported_files
        assert "excel" in exported_files
        assert "html" in exported_files
        assert "pdf" in exported_files


class TestExportPdf:
    @pytest.mark.asyncio
    async def test_export_pdf_threads_ui_config_to_pdf_report(self, monkeypatch, tmp_path):
        """
        Regression for #69: ``DataManager`` must pass its ``ui_config`` into the
        ``PDFReport`` it constructs. Before the fix, ``_export_pdf`` did
        ``PDFReport(date_format=date_format)`` with no ``ui_config``, so the PDF
        fell through to :class:`UIDefaults` placeholders ("Default header text"
        / "Default footer text") regardless of what the user had configured.
        """
        captured: dict = {}

        def fake_pdf_report(date_format, ui_config=None):
            captured["date_format"] = date_format
            captured["ui_config"] = ui_config
            # Minimal stand-in for the PDFReport surface ``_export_pdf`` exercises.
            # MagicMock auto-creates ``add_page``/``add_table_header``/``set_font``
            # /``cell``/``output``; we only need to hand-set the attributes that
            # participate in arithmetic or iteration.
            pdf = MagicMock()
            pdf.ui_config = ui_config or {}
            pdf.calculate_column_widths = lambda df: [10] * len(df.columns)
            pdf.h = 200
            pdf.get_y = lambda: 0
            return pdf

        monkeypatch.setattr("src.patcher.core.data_manager.PDFReport", fake_pdf_report)

        custom_ui = {
            "header_text": "May-Patch-Report-2026",
            "footer_text": "Confidential",
            "font_name": "Helvetica",
            "reg_font_path": "",
            "bold_font_path": "",
            "logo_path": "",
            "header_color": "#abcdef",
        }
        data_manager = DataManager(disable_cache=True, ui_config=custom_ui)
        df = pd.DataFrame({"Title": ["Firefox"], "Latest Version": ["121.0"]})
        pdf_path = tmp_path / "test.pdf"

        await data_manager._export_pdf(df, pdf_path, "%Y-%m-%d")

        # The fix: ``ui_config`` reaches ``PDFReport`` so the configured header
        # text actually renders, instead of the ``UIDefaults`` placeholder.
        assert captured["ui_config"] is custom_ui
        assert captured["ui_config"]["header_text"] == "May-Patch-Report-2026"
        assert captured["ui_config"]["footer_text"] == "Confidential"
