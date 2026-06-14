"""Multi-format rendering of patch titles to PDF, Excel, HTML, and JSON reports."""

import asyncio
import json
import re
from datetime import datetime
from pathlib import Path
from string import Template

import pandas as pd
from fpdf.enums import XPos, YPos
from openpyxl.utils import get_column_letter

from ..policy import IGNORED_EXPORT_COLUMNS
from .exceptions import PatcherError
from .logger import LogMe
from .models.patch import PatchDevice, PatchTitle
from .pdf_report import PDFReport
from .serialization import titles_to_dict


class Exporter:
    """Renders a list of ``PatchTitle`` objects to report files (PDF, Excel, HTML, JSON)."""

    def __init__(self, patch_titles: list[PatchTitle], ui_config: dict | None = None) -> None:
        """
        The ``Exporter`` renders patch titles to report files. It is a pure
        consumer: the canonical DataFrame is built and cached upstream by
        :class:`~patcher.core.data_manager.DataManager` and handed in to
        :meth:`export`; the exporter never touches the cache.

        :param patch_titles: Titles backing the report (used for the JSON payload).
        :type patch_titles: list[:class:`~patcher.core.models.patch.PatchTitle`]
        :param ui_config: Optional dict of UI settings (header text, footer
            text, font paths, logo, header color) forwarded to
            :class:`PDFReport` when generating PDF output. When ``None``,
            ``PDFReport`` falls back to :class:`UIDefaults` values.
        :type ui_config: dict | None
        """
        self.patch_titles = patch_titles
        self.ui_config = ui_config
        self.log = LogMe(self.__class__.__name__)

    def serialize_titles_to_dict(self, report_title: str | None = None) -> dict:
        """Serialize the patch titles to a JSON-ready dict (see :func:`~patcher.core.serialization.titles_to_dict`)."""
        return titles_to_dict(self.patch_titles, report_title=report_title)

    @staticmethod
    def _generate_filename(output_dir: str | Path, extension: str, analysis: bool = False) -> Path:
        """Formats naming of exported reports based upon context (analyze/export)."""
        output_dir = Path(output_dir)
        current_date = datetime.now().strftime("%m-%d-%y")

        if analysis and extension != "html":
            raise PatcherError("Only HTML format is supported for analysis.", received=extension)

        if analysis:
            export_dir = output_dir / "Patch-Analysis-Reports"
            filename = f"patch-analysis-{current_date}.{extension}"
        else:
            export_dir = output_dir
            filename = f"patch-report-{current_date}.{extension}"

        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir / filename

    async def _export_json(self, json_path: Path, report_title: str | None) -> None:
        """Writes the serialized titles to ``json_path``."""
        payload = self.serialize_titles_to_dict(report_title=report_title)
        await asyncio.to_thread(json_path.write_text, json.dumps(payload, indent=2))

    async def _export_pdf(self, df: pd.DataFrame, pdf_path: Path, date_format: str):
        """Generates a PDF Report from a given DataFrame."""
        pdf = PDFReport(date_format=date_format, ui_config=self.ui_config)
        pdf.table_headers = df.columns.tolist()
        pdf.column_widths = pdf.calculate_column_widths(df)

        pdf.add_page()
        pdf.add_table_header()
        pdf.set_font(pdf.ui_config.get("font_name"), "", 9)

        # Data rows
        for _, row in df.iterrows():
            for data, width in zip(row.astype(str), pdf.column_widths):
                pdf.cell(width, 10, str(data), border=1, align="C")
            pdf.cell(0, 10, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            if pdf.get_y() > pdf.h - 20:
                pdf.add_page()
                pdf.add_table_header()

        # Save PDF to a file
        try:
            await asyncio.to_thread(lambda: pdf.output(str(pdf_path)))
            self.log.info(f"PDF report created as expected: {pdf_path}")
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Unable to export PDF report.",
                file_path=pdf_path,
                error_msg=str(e),
            )

    async def _export_html(
        self,
        df: pd.DataFrame,
        html_path: Path,
        report_title: str,
        date_format: str,
        header_color: str | None,
    ):
        """Generates an HTML report from a given DataFrame."""
        if not header_color:
            from .models.settings import UIDefaults

            header_color = UIDefaults().header_color

        hover_color = self._darken_color(header_color, 0.2)  # darken hover
        # Ensure header color is properly formatted
        if not header_color.startswith("#"):
            header_color = f"#{header_color}"

        headers = "".join(
            f'<th onclick="sortTable({i})">{field.replace("_", " ").title()}</th>'
            for i, field in enumerate(df.columns)
        )
        rows = "".join(
            f"<tr>{''.join(f'<td>{cell}</td>' for cell in row)}</tr>" for row in df.values
        )

        template = Template((Path(__file__).parent.parent / "templates/analysis.html").read_text())
        rendered_html = template.substitute(
            title=report_title,
            heading=report_title,
            date=datetime.now().strftime(date_format),
            headers=headers,
            rows=rows,
            header_color=header_color,
            hover_color=hover_color,
        )

        try:
            await asyncio.to_thread(lambda: html_path.write_text(rendered_html, encoding="utf-8"))
            self.log.info(f"HTML report exported successfully to {html_path}")
        except OSError as e:
            raise PatcherError(
                "Error saving HTML file.", file_path=str(html_path), error_msg=str(e)
            )

    def _darken_color(self, hex_color: str, factor: float = 0.2) -> str:
        """Darkens a hex color by a given factor (0.0 to 1.0)."""
        if not hex_color:
            hex_color = "#6432bdff"

        hex_color = hex_color.lstrip("#")
        if len(hex_color) == 8:
            hex_color = hex_color[:6]  # strip alpha channel

        # RGB conversion
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        # Darken
        r = int(r * (1 - factor))
        g = int(g * (1 - factor))
        b = int(b * (1 - factor))

        return f"#{r:02x}{g:02x}{b:02x}"

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """
        Sanitize a string to be a valid Excel sheet name.

        Excel sheet names must be ≤31 characters and cannot contain: : \\ / ? * [ ]

        :param name: The desired sheet name.
        :type name: str
        :return: Sanitized sheet name valid for Excel.
        :rtype: str
        """
        sanitized = re.sub(r"[:\\/?*\[\]]", "", name)
        return sanitized[:31]

    def _write_multisheet_workbook(
        self,
        excel_path: Path,
        patch_data_df: pd.DataFrame,
        device_reports: dict[str, list[PatchDevice]],
    ) -> None:
        """
        Write Excel workbook with patch data and per-application device sheets.

        Creates a workbook with the main patch data (PatchTitle info) on the first
        sheet, then adds individual sheets for each application title containing
        device-level patch data.

        :param excel_path: Path where Excel file will be saved.
        :type excel_path: ~pathlib.Path
        :param patch_data_df: DataFrame containing patch title data (from PatchTitle objects).
        :type patch_data_df: pd.DataFrame
        :param device_reports: Dictionary mapping title IDs to lists of PatchDevice objects.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]]
        """
        title_lookup = {pt.title_id: pt.title for pt in self.patch_titles}

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            patch_data_df.to_excel(writer, sheet_name="Patch Report", index=False)
            self.log.debug("Added main Patch Report sheet to workbook")

            for title_id, devices in device_reports.items():
                if not devices:
                    self.log.debug(f"No devices found for title {title_id}, skipping")
                    continue

                title_name = title_lookup.get(title_id, f"Title_{title_id}")
                sheet_name = self._sanitize_sheet_name(title_name)

                device_rows = [
                    {
                        "Computer Name": device.computer_name,
                        "Device ID": device.device_id,
                        "Username": device.username,
                        "OS Version": device.operating_system_version,
                        "App Version": device.version,
                        "Last Contact": device.last_contact_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "Department": device.department_name or "",
                        "Building": device.building_name or "",
                        "Site": device.site_name or "",
                    }
                    for device in devices
                ]

                device_df = pd.DataFrame(device_rows)
                device_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Autosize columns for readability
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(device_df.columns):
                    max_len = max(device_df[col].astype(str).map(len).max(), len(col)) + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = max_len

                self.log.debug(f"Added sheet '{sheet_name}' with {len(devices)} devices")

    async def _export_excel(
        self,
        output_dir: Path,
        df: pd.DataFrame,
        device_reports: dict[str, list[PatchDevice]] | None = None,
        analysis: bool = False,
    ) -> Path:
        """
        Exports a DataFrame to an Excel file, optionally including per-title device sheets.

        :param output_dir: Directory to save the Excel file.
        :type output_dir: ~pathlib.Path
        :param df: The summary DataFrame to export.
        :type df: pd.DataFrame
        :param device_reports: Optional dictionary mapping title IDs to device lists.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]] | None
        :param analysis: Whether this is an analysis report.
        :type analysis: bool
        :return: Path to the exported Excel file.
        :rtype: ~pathlib.Path
        """
        excel_path = self._generate_filename(output_dir, "xlsx", analysis)

        try:
            if device_reports:
                await asyncio.to_thread(
                    self._write_multisheet_workbook, excel_path, df, device_reports
                )
            else:
                await asyncio.to_thread(df.to_excel, excel_path, index=False)

            self.log.info(f"Excel report created successfully to {excel_path}.")
            return excel_path
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered error saving DataFrame.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

    async def export(
        self,
        df: pd.DataFrame,
        output_dir: str | Path,
        report_title: str,
        analysis: bool = False,
        date_format: str = "%B %d %Y",
        formats: set[str] | None = None,
        header_color: str | None = None,
        device_reports: dict[str, list[PatchDevice]] | None = None,
    ) -> dict[str, str]:
        """
        Render the patch data to the specified report formats.

        :param df: The canonical, already-cached DataFrame to render. Presentation-
            excluded columns are dropped here before rendering.
        :type df: pandas.DataFrame
        :param output_dir: The directory in which to save the exported report(s).
        :type output_dir: str | ~pathlib.Path
        :param report_title: The title to use for the header in exported report(s).
        :type report_title: str
        :param analysis: Denotes whether this is an analysis report (affects HTML output path).
        :type analysis: bool
        :param date_format: The date format for PDF/HTML headers. Defaults to "%B %d %Y".
        :type date_format: str
        :param formats: A set of formats to export. Defaults to all ({"excel", "html", "pdf", "json"}).
        :type formats: set | None
        :param header_color: Hex color to use for HTML header table background.
            Falls back to :attr:`~patcher.core.models.settings.UIDefaults.header_color` when ``None``.
        :type header_color: str | None
        :param device_reports: Optional dictionary mapping title IDs to device lists for per-title detail sheets.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]] | None
        :return: A dictionary containing paths to generated reports.
        :rtype: dict[str, str]
        """
        if formats is None:
            formats = {"excel", "html", "pdf", "json"}

        exported_files = {}

        df = df.drop(
            columns=[col.replace("_", " ").title() for col in IGNORED_EXPORT_COLUMNS],
            errors="ignore",
        )

        # Verification of directory existence runs synchronously
        output_dir = Path(output_dir)

        # Excel
        if "excel" in formats:
            try:
                excel_path = await self._export_excel(output_dir, df, device_reports, analysis)
                exported_files["excel"] = str(excel_path)
            except (OSError, PermissionError) as e:
                raise PatcherError(
                    "Encountered error saving DataFrame.",
                    file_path=str(output_dir),
                    error_msg=str(e),
                )

        try:
            # PDF
            if "pdf" in formats:
                pdf_path = self._generate_filename(output_dir, "pdf", analysis)
                await self._export_pdf(df, pdf_path, date_format)
                exported_files["pdf"] = str(pdf_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving PDF report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        try:
            # HTML
            if "html" in formats:
                html_path = self._generate_filename(output_dir, "html", analysis)
                await self._export_html(df, html_path, report_title, date_format, header_color)
                exported_files["html"] = str(html_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving HTML report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        try:
            # JSON: serialized straight from PatchTitle models (no DataFrame round-trip)
            # so consumers get the same shape the library exposes programmatically.
            if "json" in formats:
                json_path = self._generate_filename(output_dir, "json", analysis)
                await self._export_json(json_path, report_title)
                exported_files["json"] = str(json_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving JSON report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        output_paths = "\n".join(list(exported_files.values()))
        self.log.info(f"Exported {len(exported_files)} reports as expected: {output_paths}")
        return exported_files
