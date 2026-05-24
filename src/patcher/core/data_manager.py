import asyncio
import json
import pickle
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from string import Template

import pandas as pd
from fpdf.enums import XPos, YPos
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from .exceptions import PatcherError
from .logger import LogMe
from .models.patch import PatchDevice, PatchTitle
from .pdf_report import PDFReport


def serialize_titles_to_dict(
    patch_titles: list[PatchTitle], report_title: str | None = None
) -> dict:
    """
    Convert a list of :class:`PatchTitle` into a JSON-serializable dict.

    Pure function with no I/O. Safe to call from CLI export, library code,
    or anywhere a structured representation of the titles is useful.

    The returned dict has the shape::

        {
            "generated_at": "2026-05-04T18:30:00+00:00",
            "report_title": "...",
            "title_count": 42,
            "titles": [<PatchTitle.model_dump()>, ...]
        }

    :param patch_titles: List of patch titles to serialize.
    :type patch_titles: list[:class:`PatchTitle`]
    :param report_title: Optional title carried through to consumers.
    :type report_titles: str | None
    :return: A dict ready for ``json.dump`` or direct programmatic use.
    """
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "report_title": report_title,
        "title_count": len(patch_titles),
        "titles": [title.model_dump(mode="json") for title in patch_titles],
    }


class DataManager:
    _IGNORED = ["install_label", "homebrew_cask", "title_id"]

    def __init__(self, disable_cache: bool = False):
        """
        The ``DataManager`` class handles data management for patch reports, including caching, validation and exporting to Excel.

        Data caching can be disabled by setting ``disable_cache`` to ``True`` at runtime.

        :param disable_cache: Whether caching functionality should be disabled.
        :type disable_cache: bool
        """
        self.cache_dir = Path.home() / "Library/Caches/Patcher"
        self.cache_expiration_days = 90  # Increase for better trend analysis
        self.latest_excel_file: Path | None = None
        self.log = LogMe(self.__class__.__name__)
        self._disabled = disable_cache
        self._titles: list[PatchTitle] | None = None

    @property
    def cache_off(self) -> bool:
        """
        Indicates whether caching is disabled.

        :return: True if caching is disabled, False otherwise.
        :rtype: bool
        """
        return self._disabled

    @property
    def titles(self) -> list[PatchTitle]:
        """
        Retrieve and validate the current list of ``PatchTitle`` objects.

        If titles are not already loaded, they are fetched from the latest available dataset.

        :return: The validated list of ``PatchTitle`` objects.
        :rtype: list[:class:`~patcher.core.models.patch.PatchTitle`]
        :raises PatcherError: If validation fails.
        """
        if self._titles is None:
            self.log.debug("Attempting to load PatchTitle objects from the latest dataset.")
            df = self._validate_data()
            self._titles = self._create_patches(df)
        return self._titles

    @titles.setter
    def titles(self, value: list[PatchTitle]) -> None:
        """
        Validates and sets the PatchTitle objects. Ensures the list is non-empty.

        :param value: The list of PatchTitle objects to validate.
        :type value: list[PatchTitle]
        :raises PatcherError: If value is not an iterable object.
        :raises PatcherError: If any object in the passed iterable object is not a ``PatchTitle`` object, or if titles could not be validated.
        """
        if not isinstance(value, list):
            raise PatcherError(f"Value {value} must be an list of PatchTitle objects.")

        validated_titles = []
        for item in value:
            if not isinstance(item, PatchTitle):
                raise PatcherError(f"Item {item} in list is not of PatchTitle type.")
            validated_titles.append(item)

        if not validated_titles:  # Ensure the list is not empty
            raise PatcherError("PatchTitles cannot be set to an empty list.")

        self._titles = validated_titles

    def _validate_data(self) -> pd.DataFrame:
        dataset = self.get_latest_dataset()
        if not dataset:
            raise PatcherError("No dataset available, unable to proceed with validation.")
        try:
            if dataset.suffix == ".pkl":
                with open(dataset, "rb") as f:
                    return pickle.load(f)
            elif dataset.suffix in [".xlsx", ".xls"]:
                return pd.read_excel(dataset)
            else:
                raise PatcherError("Unsupported data format for dataset.", received=dataset.suffix)
        except (FileNotFoundError, pd.errors.EmptyDataError, pickle.UnpicklingError) as e:
            raise PatcherError("Error encountered validating dataset.", error_msg=str(e))

    def _cache_data(self, df: pd.DataFrame):
        """Cache exported data for later use."""
        if self.cache_off:
            return  # Only cache if enabled

        timestamp = datetime.now().strftime("%Y%m%d%I%M")
        cache_file = self.cache_dir / f"patch_data_{timestamp}.pkl"
        self.log.debug(f"Attempting to cache data to {cache_file}.")
        try:
            with open(cache_file, "wb") as file:
                pickle.dump(df, file)  # type: ignore
            self.log.info(f"Cached data successfully to {cache_file}")
            self._clean_cache()
        except (FileNotFoundError, pickle.PicklingError, PermissionError, OSError) as e:
            exception_name = type(e).__name__
            self.log.warning(
                f"Unable to cache data to {cache_file} as expected due to {exception_name}. Details: {e}"
            )
            return

    def _clean_cache(self):
        """Remove cache files older than expiration policy."""
        expiration_time = datetime.now() - timedelta(days=self.cache_expiration_days)
        self.log.debug(f"Attempting to remove cache files older than {expiration_time}.")
        for file in self.cache_dir.iterdir():
            if file.is_file() and file.suffix == ".pkl":
                file_time = datetime.fromtimestamp(file.stat().st_mtime)
                if file_time < expiration_time:
                    try:
                        file.unlink()
                        self.log.info(f"Deleted expired cache file: {file}")
                    except OSError as e:
                        self.log.warning(f"Failed to delete cache file {file}. Details: {e}")
                        return

    def _create_dataframe(self, patch_titles: list[PatchTitle]) -> pd.DataFrame:
        """Converts list of PatchTitles into a pandas DataFrame."""
        self.log.debug("Attempting to create DataFrame from PatchTitle objects.")
        try:
            df = pd.DataFrame([patch.model_dump() for patch in patch_titles])
            df.columns = [column.replace("_", " ").title() for column in df.columns]
            self.log.info(
                f"Created DataFrame from {len(patch_titles)} PatchTitle objects successfully."
            )
            return df
        except (ValueError, pd.errors.EmptyDataError) as e:
            raise PatcherError("Encountered error creating DataFrame.", error_msg=str(e))

    def _create_patches(self, df: pd.DataFrame) -> list[PatchTitle]:
        """Convert a pandas DataFrame into a list of PatchTitle objects."""
        self.log.debug(f"Creating PatchTitle objects from DataFrame with {len(df)} rows.")
        skipped_rows = 0
        patch_titles = []

        for _, row in df.iterrows():
            try:
                patch = PatchTitle(
                    **{str(key).lower().replace(" ", "_"): value for key, value in row.items()}
                )
                patch_titles.append(patch)
            except (KeyError, ValueError, TypeError, ValidationError) as e:
                exception_name = type(e).__name__
                self.log.warning(
                    f"Encountered {exception_name} during PatchTitle creation. Skipping row. Details: {e}."
                )
                skipped_rows += 1

        if skipped_rows > 0:
            self.log.warning(f"{skipped_rows} rows were skipped during PatchTitle creation.")
        self.log.info(f"Successfully created {len(patch_titles)} PatchTitle objects.")

        return patch_titles

    @staticmethod
    def _generate_filename(output_dir: str | Path, extension: str, analysis: bool = False) -> Path:
        """Formats naming of exported HTML reports based upon context (analyze/export)."""
        output_dir = Path(output_dir)
        current_date = datetime.now().strftime("%m-%d-%y")

        if analysis and extension != "html":
            raise PatcherError("Only HTML format is supported for analysis.", received=extension)

        if analysis:
            export_dir = output_dir / "Patch-Analysis-Reports"
            filename = f"patch-analysis-{current_date}.{extension}"
        else:
            export_dir = output_dir
            filename = f"patch-report-{current_date}.{extension}"

        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir / filename

    async def _export_json(
        self,
        patch_titles: list[PatchTitle],
        json_path: Path,
        report_title: str | None,
    ) -> None:
        """Writes the serialized titles to ``json_path``."""
        payload = serialize_titles_to_dict(patch_titles, report_title=report_title)
        await asyncio.to_thread(json_path.write_text, json.dumps(payload, indent=2))

    async def _export_pdf(self, df: pd.DataFrame, pdf_path: Path, date_format: str):
        """Generates a PDF Report from a given DataFrame."""
        pdf = PDFReport(date_format=date_format)
        pdf.table_headers = df.columns.tolist()
        pdf.column_widths = pdf.calculate_column_widths(df)

        pdf.add_page()
        pdf.add_table_header()
        pdf.set_font(pdf.ui_config.get("font_name"), "", 9)

        # Data rows
        for _, row in df.iterrows():
            for data, width in zip(row.astype(str), pdf.column_widths):
                pdf.cell(width, 10, str(data), border=1, align="C")
            pdf.cell(0, 10, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            if pdf.get_y() > pdf.h - 20:
                pdf.add_page()
                pdf.add_table_header()

        # Save PDF to a file
        try:
            await asyncio.to_thread(lambda: pdf.output(str(pdf_path)))
            self.log.info(f"PDF report created as expected: {pdf_path}")
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Unable to export PDF report.",
                file_path=pdf_path,
                error_msg=str(e),
            )

    async def _export_html(
        self,
        df: pd.DataFrame,
        html_path: Path,
        report_title: str,
        date_format: str,
        header_color: str | None,
    ):
        """Generates an HTML report from a given DataFrame."""
        if not header_color:
            from .models.ui import UIDefaults

            header_color = UIDefaults().header_color

        hover_color = self._darken_color(header_color, 0.2)  # darken hover
        # Ensure header color is properly formatted
        if not header_color.startswith("#"):
            header_color = f"#{header_color}"

        headers = "".join(
            f'<th onclick="sortTable({i})">{field.replace("_", " ").title()}</th>'
            for i, field in enumerate(df.columns)
        )
        rows = "".join(
            f"<tr>{''.join(f'<td>{cell}</td>' for cell in row)}</tr>" for row in df.values
        )

        template = Template((Path(__file__).parent.parent / "templates/analysis.html").read_text())
        rendered_html = template.substitute(
            title=report_title,
            heading=report_title,
            date=datetime.now().strftime(date_format),
            headers=headers,
            rows=rows,
            header_color=header_color,
            hover_color=hover_color,
        )

        try:
            await asyncio.to_thread(lambda: html_path.write_text(rendered_html, encoding="utf-8"))
            self.log.info(f"HTML report exported successfully to {html_path}")
        except OSError as e:
            raise PatcherError(
                "Error saving HTML file.", file_path=str(html_path), error_msg=str(e)
            )

    def _darken_color(self, hex_color: str, factor: float = 0.2) -> str:
        """Darkens a hex color by a given factor (0.0 to 1.0)."""
        if not hex_color:
            hex_color = "#6432bdff"

        hex_color = hex_color.lstrip("#")
        if len(hex_color) == 8:
            hex_color = hex_color[:6]  # strip alpha channel

        # RGB conversion
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        # Darken
        r = int(r * (1 - factor))
        g = int(g * (1 - factor))
        b = int(b * (1 - factor))

        return f"#{r:02x}{g:02x}{b:02x}"

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """
        Sanitize a string to be a valid Excel sheet name.

        Excel sheet names must be ≤31 characters and cannot contain: : \\ / ? * [ ]

        :param name: The desired sheet name.
        :type name: str
        :return: Sanitized sheet name valid for Excel.
        :rtype: str
        """
        sanitized = re.sub(r"[:\\/?*\[\]]", "", name)
        return sanitized[:31] if len(sanitized) > 31 else sanitized

    def _write_multisheet_workbook(
        self,
        excel_path: Path,
        patch_data_df: pd.DataFrame,
        patch_titles: list[PatchTitle],
        device_reports: dict[str, list[PatchDevice]],
    ) -> None:
        """
        Write Excel workbook with patch data and per-application device sheets.

        Creates a workbook with the main patch data (PatchTitle info) on the first
        sheet, then adds individual sheets for each application title containing
        device-level patch data.

        :param excel_path: Path where Excel file will be saved.
        :type excel_path: ~pathlib.Path
        :param patch_data_df: DataFrame containing patch title data (from PatchTitle objects).
        :type patch_data_df: pd.DataFrame
        :param device_reports: Dictionary mapping title IDs to lists of PatchDevice objects.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]]
        """
        title_lookup = {pt.title_id: pt.title for pt in patch_titles}

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            patch_data_df.to_excel(writer, sheet_name="Patch Report", index=False)
            self.log.debug("Added main Patch Report sheet to workbook")

            for title_id, devices in device_reports.items():
                if not devices:
                    self.log.debug(f"No devices found for title {title_id}, skipping")
                    continue

                title_name = title_lookup.get(title_id, f"Title_{title_id}")
                sheet_name = self._sanitize_sheet_name(title_name)

                device_rows = [
                    {
                        "Computer Name": device.computer_name,
                        "Device ID": device.device_id,
                        "Username": device.username,
                        "OS Version": device.operating_system_version,
                        "App Version": device.version,
                        "Last Contact": device.last_contact_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "Department": device.department_name or "",
                        "Building": device.building_name or "",
                        "Site": device.site_name or "",
                    }
                    for device in devices
                ]

                device_df = pd.DataFrame(device_rows)
                device_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Autosize columns for readability
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(device_df.columns):
                    max_len = max(device_df[col].astype(str).map(len).max(), len(col)) + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = max_len

                self.log.debug(f"Added sheet '{sheet_name}' with {len(devices)} devices")

    async def _export_excel(
        self,
        output_dir: Path,
        df: pd.DataFrame,
        patch_titles: list[PatchTitle] | None = None,
        device_reports: dict[str, list[PatchDevice]] | None = None,
        analysis: bool = False,
    ) -> Path:
        """
        Exports a DataFrame to an Excel file, optionally including per-title device sheets.

        :param output_dir: Directory to save the Excel file.
        :type output_dir: ~pathlib.Path
        :param df: The summary DataFrame to export.
        :type df: pd.DataFrame
        :param analysis: Whether this is an analysis report.
        :type analysis: bool
        :param patch_titles:
        :type patch_titles: list[:class:`~patcher.core.models.patch.PatchTitle`]
        :param device_reports: Optional dictionary mapping title IDs to device lists.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]] | None
        :return: Path to the exported Excel file.
        :rtype: ~pathlib.Path
        """
        excel_path = self._generate_filename(output_dir, "xlsx", analysis)

        try:
            if device_reports and patch_titles:
                await asyncio.to_thread(
                    self._write_multisheet_workbook, excel_path, df, patch_titles, device_reports
                )
            else:
                await asyncio.to_thread(df.to_excel, excel_path, index=False)

            self.latest_excel_file = excel_path
            self.log.info(f"Excel report created successfully to {excel_path}.")
            return excel_path
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered error saving DataFrame.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

    async def export(
        self,
        patch_titles: list[PatchTitle],
        output_dir: str | Path,
        report_title: str,
        analysis: bool = False,
        date_format: str = "%B %d %Y",
        formats: set[str] | None = None,
        header_color: str | None = None,
        device_reports: dict[str, list[PatchDevice]] | None = None,
    ) -> dict[str, str]:
        """
        Exports patch data to the specified formats.

        :param patch_titles: A list of ``PatchTitle`` objects to include in the report. Defaults to ``self.titles`` if not provided
        :type patch_titles: list [:class:`~patcher.core.models.patch.PatchTitle`]
        :param output_dir: The directory in which to save the exported report(s).
        :type output_dir: str | ~pathlib.Path
        :param report_title: The title to use for the header in exported report(s). Defaults to the ``header_text`` key in ``com.liquidzoo.patcher.plist``.
        :type report_title: str
        :param analysis: Denotes whether this is analysis report (affects HTML output path).
        :type analysis: bool
        :param date_format: The date format for PDF/HTML headers. Defaults to "%B %d %Y" (Month Day Year).
        :type date_format: str
        :param formats: A set of formats to export. Defaults to all ({"excel", "html", "pdf"}).
        :type formats: set | None
        :param header_color: Hex color to use for HTML header table background.
            Falls back to :attr:`~patcher.core.models.ui.UIDefaults.header_color` when ``None``.
        :type header_color: str | None
        :param device_reports: Optional dictionary mapping title IDs to device lists for per-title detail sheets.
        :type device_reports: dict[str, list[:class:`~patcher.core.models.patch.PatchDevice`]] | None
        :return: A dictionary containing paths to generated reports.
        :rtype: dict[str, str]
        """
        if formats is None:
            formats = {"excel", "html", "pdf", "json"}

        exported_files = {}

        patch_titles = patch_titles or self.titles
        df = await asyncio.to_thread(self._create_dataframe, patch_titles)
        await asyncio.to_thread(self._cache_data, df)

        df = df.drop(
            columns=[col.replace("_", " ").title() for col in DataManager._IGNORED], errors="ignore"
        )

        # Surface Homebrew Cask coverage as a readable column. The raw
        # list-of-dicts ``homebrew_cask`` field is dropped via ``_IGNORED``
        # (same treatment as ``install_label``); this derived column shows the
        # matched cask token(s) instead. Positionally aligned with
        # ``patch_titles`` since ``_create_dataframe`` preserves their order.
        # Added only when at least one title matched a cask, so default
        # (Installomator-only) exports are unchanged.
        if any(title.homebrew_cask for title in patch_titles):
            df["Homebrew"] = [
                ", ".join(match.token for match in (title.homebrew_cask or []))
                for title in patch_titles
            ]

        # Verification of directory existence runs synchronously
        output_dir = Path(output_dir)

        # Excel
        if "excel" in formats:
            try:
                excel_path = await self._export_excel(
                    output_dir, df, patch_titles, device_reports, analysis
                )
                self.latest_excel_file = excel_path
                exported_files["excel"] = str(excel_path)
            except (OSError, PermissionError) as e:
                raise PatcherError(
                    "Encountered error saving DataFrame.",
                    file_path=str(output_dir),
                    error_msg=str(e),
                )

        try:
            # PDF
            if "pdf" in formats:
                pdf_path = self._generate_filename(output_dir, "pdf", analysis)
                await self._export_pdf(df, pdf_path, date_format)
                exported_files["pdf"] = str(pdf_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving PDF report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        try:
            # HTML
            if "html" in formats:
                html_path = self._generate_filename(output_dir, "html", analysis)
                await self._export_html(df, html_path, report_title, date_format, header_color)
                exported_files["html"] = str(html_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving HTML report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        try:
            # JSON: serialized straight from PatchTitle models (no DataFrame round-trip)
            # so consumers get the same shape the library exposes programmatically.
            if "json" in formats:
                json_path = self._generate_filename(output_dir, "json", analysis)
                await self._export_json(patch_titles, json_path, report_title)
                exported_files["json"] = str(json_path)
        except (OSError, PermissionError) as e:
            raise PatcherError(
                "Encountered an error saving JSON report.",
                file_path=str(output_dir),
                error_msg=str(e),
            )

        output_paths = "\n".join(list(exported_files.values()))
        self.log.info(f"Exported {len(exported_files)} reports as expected: {output_paths}")
        return exported_files

    def reset_cache(self) -> bool:
        """
        Removes all cached files from Cache directory. See :ref:`reset <resetting_patcher>`.

        :return: True if all files were able to be removed, False otherwise.
        :rtype: bool
        """
        try:
            [file.unlink() for file in self.get_cached_files()]
            return True
        except OSError as e:
            self.log.warning(f"Encountered {type(e).__name__} during cache reset. Details: {e}")
            return False

    def load_cached_data(self) -> list[pd.DataFrame]:
        """
        Load all cached data files into a list of DataFrames.

        :return: list of pandas DataFrame objects with cached data.
        :rtype: list[~pandas.DataFrame]
        """
        dataframes = []
        cached_files = self.get_cached_files()
        for file in cached_files:
            try:
                with open(file, "rb") as f:
                    dataframes.append(pickle.load(f))
                self.log.info(f"Loaded cache data from {file}")
            except (pickle.UnpicklingError, FileNotFoundError) as e:
                self.log.warning(f"Failed to load cached file {file}. Details: {e}")
        return dataframes

    def get_cached_files(self) -> list[Path]:
        """
        Retrieves all cached file Paths.

        :return: A list of ``Path`` objects pointing to cached files.
        :rtype: list[~pathlib.Path]
        """
        return [file for file in self.cache_dir.iterdir() if file.suffix == ".pkl"]

    def get_latest_dataset(self) -> Path | None:
        """
        Retrieves the most recent dataset of patch reports and returns the path.

        If a tracked Excel file is available, it is preferred. Otherwise, searches the cache directory.

        :return: Path to the latest dataset (Excel or pickle file), or None if no dataset is found.
        :rtype: ~pathlib.Path | None
        """
        if self.latest_excel_file and self.latest_excel_file.exists():
            self.log.info(f"Using latest tracked Excel file: {self.latest_excel_file}")
            return self.latest_excel_file

        pickle_files = sorted(
            self.cache_dir.glob("*.pkl"), key=lambda f: f.stat().st_mtime, reverse=True
        )
        if pickle_files:
            self.log.info(f"Using latest cached pickle file: {pickle_files[0]}")
            return pickle_files[0]

        self.log.warning("No datasets found (Excel or pickle).")
        return None
